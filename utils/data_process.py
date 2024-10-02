import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os


def generate_array():
    array = np.empty((384, 5))
    for i in range(24):
        for j in range(16):
            new_data = [i + 1, -0.5, 0, 0.866025, j ]
            array[i * 16 + j] = new_data
    return array

def process_and_append_data(folder_path, output_path):
    # 获取所有文件名并按照自然顺序排序
    file_list = sorted([f for f in os.listdir(folder_path) if f.endswith(".dat")])

    # 用于存储所有文件的数据
    all_data = []

    # 遍历文件夹中的所有.dat文件
    for i, filename in enumerate(file_list):
        # 构造完整的文件路径
        file_path = os.path.join(folder_path, filename)
        
        # 读取文件并跳过第一行的列名（因为没有合适的分隔符，我们手动命名列）
        col_names = ['Wavelength', 'CH1', 'CH2', 'CH3', 'CH4', 'CH5', 'CH6', 'CH7', 'CH8']
        data = pd.read_csv(file_path, sep='\s+', names=col_names, skiprows=2)

        # 提取 Wavelength、CH2 和 CH3 列，并重命名 CH2 和 CH3 列为独特的名字
        df = data[['Wavelength', 'CH2', 'CH3']].copy()
        df.columns = ['Wavelength', f'CH2_File_{i+1}', f'CH3_File_{i+1}']
        
        all_data.append(df)

    # 使用 pd.concat 一次性合并所有数据
    new_data = pd.concat(all_data, axis=1)
    new_data = new_data.loc[:, ~new_data.columns.duplicated()]  # 移除重复的 Wavelength 列

    # 检查输出文件是否存在
    if os.path.exists(output_path):
        # 如果文件存在，读取现有数据
        existing_data = pd.read_excel(output_path)
        
        # 合并现有数据和新数据
        merged_data = pd.concat([existing_data, new_data.iloc[:, 1:]], axis=1)
    else:
        # 如果文件不存在，使用新数据创建
        merged_data = new_data

    # 将结果写入Excel文件
    merged_data.to_excel(output_path, index=False)
    print(f"数据已成功添加到 {output_path}")


def append_to_excel(array, filename, sheet_name):
    # 如果文件不存在，创建新文件
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    else:
        wb = openpyxl.load_workbook(filename)
    
    # 如果指定的工作表不存在，创建新的工作表
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    # 创建 DataFrame
    df = pd.DataFrame(y, columns=['Position', 'Column2', 'Column3', 'Column4', 'Column5'])
    
    # 获取工作表中最后一行的行号
    last_row = ws.max_row
    
    # 将DataFrame写入工作表，从最后一行之后开始
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, value in enumerate(row):
            ws.cell(row=last_row + 1 + r_idx, column=c_idx + 1, value=value)
    
    # 保存Excel文件
    wb.save(filename)
    print(f"数据已成功添加到 {filename} 的 {sheet_name} 工作表中")


# 主程序
if __name__ == "__main__":
    filename = '../Data_Sets/7_points_label.xlsx'  # 替换为您的Excel文件名
    sheet_name = 'sheet1'  # 新工作表的名称
    
    # 生成数组
    y = generate_array()
    
    # 将数组添加到Excel文件
    append_to_excel(y, filename, sheet_name)
    
    # # 验证：打印Excel文件中的工作表信息
    # wb = openpyxl.load_workbook(filename)
    # ws = wb[sheet_name]
    # print(f"\n工作表 '{sheet_name}' 中的数据行数: {ws.max_row}")


    # 文件夹路径
    folder_path = '../Data_Sets/Fiber_7points/bei30/Spectrum Data-20240930'  # 修改为实际路径

    # 输出Excel文件路径
    output_path = '../Data_Sets/7_points.xlsx'  # 修改为实际输出路径

    # 处理数据并追加到Excel文件
    process_and_append_data(folder_path, output_path)

    # # 验证：打印Excel文件中的列数
    # df = pd.read_excel(output_path)
    # print(f"\nExcel文件中的列数: {len(df.columns)}")
    # print(f"最后两列的名称: {df.columns[-2:]}")